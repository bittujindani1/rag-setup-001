from __future__ import annotations

import csv
import io
import json
import tempfile
import time
from collections import Counter
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable

import boto3
from botocore.exceptions import ClientError
from openpyxl import load_workbook


def _safe_dataset_id(value: str) -> str:
    normalized = (value or "").strip().lower().replace("-", "_")
    return "".join(char if char.isalnum() or char == "_" else "_" for char in normalized)


class AnalyticsStore:
    def __init__(
        self,
        *,
        region_name: str,
        bucket_name: str,
        glue_database: str,
        athena_workgroup: str = "primary",
        metrics_ttl_seconds: int = 3600,
    ) -> None:
        self.region_name = region_name
        self.bucket_name = bucket_name
        self.glue_database = glue_database
        self.athena_workgroup = athena_workgroup
        self.metrics_ttl_seconds = metrics_ttl_seconds
        self.s3 = boto3.client("s3", region_name=region_name)
        self.glue = boto3.client("glue", region_name=region_name)
        self.athena = boto3.client("athena", region_name=region_name)

    def parse_structured_file(self, file_name: str, file_bytes: bytes) -> list[dict[str, Any]]:
        suffix = Path(file_name).suffix.lower()
        if suffix == ".csv":
            text_stream = io.StringIO(file_bytes.decode("utf-8", errors="ignore"))
            return [dict(row) for row in csv.DictReader(text_stream)]
        if suffix == ".json":
            payload = json.loads(file_bytes.decode("utf-8", errors="ignore"))
            if isinstance(payload, dict):
                payload = payload.get("rows") or payload.get("items") or payload.get("tickets") or [payload]
            return [dict(row) for row in payload if isinstance(row, dict)]
        if suffix == ".xlsx":
            workbook = load_workbook(io.BytesIO(file_bytes), read_only=True, data_only=True)
            worksheet = workbook.worksheets[0]
            rows = list(worksheet.iter_rows(values_only=True))
            if not rows:
                return []
            headers = [str(cell).strip() if cell is not None else "" for cell in rows[0]]
            records: list[dict[str, Any]] = []
            for row in rows[1:]:
                record = {}
                for header, value in zip(headers, row):
                    if not header:
                        continue
                    record[header] = value
                if any(value not in (None, "") for value in record.values()):
                    records.append(record)
            return records
        raise ValueError("Unsupported analytics file type. Allowed: csv, json, xlsx.")

    def profile_schema(self, rows: list[dict[str, Any]]) -> dict[str, Any]:
        if not rows:
            return {"columns": []}
        columns = []
        for column_name in rows[0].keys():
            values = [row.get(column_name) for row in rows if row.get(column_name) not in (None, "")]
            unique_values = list(dict.fromkeys(str(value) for value in values if value is not None))
            kind = self._infer_kind(values, unique_values)
            columns.append(
                {
                    "name": column_name,
                    "kind": kind,
                    "cardinality": len(unique_values),
                    "sample_values": unique_values[:10],
                }
            )
        return {"columns": columns}

    def store_dataset(
        self,
        *,
        dataset_id: str,
        source_name: str,
        file_bytes: bytes,
        rows: list[dict[str, Any]],
        schema_profile: dict[str, Any],
    ) -> dict[str, str]:
        safe_dataset_id = _safe_dataset_id(dataset_id)
        raw_key = f"datasets/{safe_dataset_id}/raw/{source_name}"
        schema_key = f"datasets/{safe_dataset_id}/schema.json"
        data_key = f"datasets/{safe_dataset_id}/tabular/data.csv"

        self.s3.put_object(Bucket=self.bucket_name, Key=raw_key, Body=file_bytes)
        self.s3.put_object(
            Bucket=self.bucket_name,
            Key=schema_key,
            Body=json.dumps(schema_profile, indent=2).encode("utf-8"),
            ContentType="application/json",
        )
        csv_bytes = self._to_csv_bytes(rows, safe_dataset_id, schema_profile)
        self.s3.put_object(
            Bucket=self.bucket_name,
            Key=data_key,
            Body=csv_bytes,
            ContentType="text/csv",
        )
        self._ensure_glue_database()
        self._upsert_glue_table(safe_dataset_id, schema_profile, data_key)
        self._write_dataset_manifest(safe_dataset_id, source_name, schema_profile)
        return {
            "dataset_id": safe_dataset_id,
            "table_name": f"dataset_{safe_dataset_id}",
            "raw_key": raw_key,
            "data_key": data_key,
        }

    def cache_metrics(self, dataset_id: str, summary: dict[str, Any], metrics: list[dict[str, Any]]) -> None:
        payload = {
            "dataset_id": dataset_id,
            "generated_at": int(time.time()),
            "expires_at": int(time.time()) + self.metrics_ttl_seconds,
            "summary": summary,
            "metrics": metrics,
        }
        self.s3.put_object(
            Bucket=self.bucket_name,
            Key=f"datasets/{dataset_id}/metrics/summary.json",
            Body=json.dumps(payload, indent=2).encode("utf-8"),
            ContentType="application/json",
        )

    def load_metrics_cache(self, dataset_id: str) -> dict[str, Any] | None:
        try:
            response = self.s3.get_object(Bucket=self.bucket_name, Key=f"datasets/{dataset_id}/metrics/summary.json")
        except ClientError:
            return None
        payload = json.loads(response["Body"].read().decode("utf-8"))
        if payload.get("expires_at", 0) < int(time.time()):
            return None
        return payload

    def cache_metric_result(self, dataset_id: str, metric_id: str, payload: dict[str, Any]) -> None:
        wrapped_payload = {
            **payload,
            "generated_at": int(time.time()),
            "expires_at": int(time.time()) + self.metrics_ttl_seconds,
        }
        self.s3.put_object(
            Bucket=self.bucket_name,
            Key=f"datasets/{dataset_id}/metrics/results/{metric_id}.json",
            Body=json.dumps(wrapped_payload, indent=2).encode("utf-8"),
            ContentType="application/json",
        )

    def load_metric_result(self, dataset_id: str, metric_id: str) -> dict[str, Any] | None:
        try:
            response = self.s3.get_object(Bucket=self.bucket_name, Key=f"datasets/{dataset_id}/metrics/results/{metric_id}.json")
        except ClientError:
            return None
        payload = json.loads(response["Body"].read().decode("utf-8"))
        if payload.get("expires_at", 0) < int(time.time()):
            return None
        return payload

    def execute_query(self, *, dataset_id: str, sql: str) -> dict[str, Any]:
        output_location = f"s3://{self.bucket_name}/datasets/{dataset_id}/query-results/"
        execution = self.athena.start_query_execution(
            QueryString=sql,
            QueryExecutionContext={"Database": self.glue_database},
            ResultConfiguration={"OutputLocation": output_location},
            WorkGroup=self.athena_workgroup,
        )
        execution_id = execution["QueryExecutionId"]
        for _ in range(30):
            status = self.athena.get_query_execution(QueryExecutionId=execution_id)["QueryExecution"]["Status"]["State"]
            if status in {"SUCCEEDED", "FAILED", "CANCELLED"}:
                break
            time.sleep(1)
        if status != "SUCCEEDED":
            raise RuntimeError(f"Athena query failed with status {status}.")

        results = self.athena.get_query_results(QueryExecutionId=execution_id)
        rows = results["ResultSet"]["Rows"]
        headers = [item.get("VarCharValue", "") for item in rows[0]["Data"]] if rows else []
        records = []
        for row in rows[1:]:
            values = [item.get("VarCharValue") for item in row["Data"]]
            records.append({header: value for header, value in zip(headers, values)})

        return {
            "execution_id": execution_id,
            "columns": headers,
            "rows": records,
            "source": "athena",
        }

    def list_datasets(self) -> list[dict[str, Any]]:
        prefixes = self._list_common_prefixes("datasets/")
        datasets = []
        for prefix in prefixes:
            dataset_id = prefix.split("/")[-2]
            manifest = self._load_json(f"datasets/{dataset_id}/manifest.json") or {"dataset_id": dataset_id}
            datasets.append(manifest)
        return sorted(datasets, key=lambda item: item.get("updated_at", 0), reverse=True)

    def get_schema(self, dataset_id: str) -> dict[str, Any] | None:
        return self._load_json(f"datasets/{dataset_id}/schema.json")

    def get_table_name(self, dataset_id: str) -> str:
        return f"dataset_{_safe_dataset_id(dataset_id)}"

    def build_summary_metrics(self, rows: list[dict[str, Any]], schema_profile: dict[str, Any]) -> dict[str, Any]:
        total_rows = len(rows)
        summary: dict[str, Any] = {"total_rows": total_rows}
        for column in schema_profile.get("columns", []):
            if column.get("kind") != "categorical":
                continue
            name = column["name"]
            counts = Counter(str(row.get(name, "unknown")) for row in rows if row.get(name) not in (None, ""))
            if counts:
                summary[f"top_{name}"] = {
                    "label": counts.most_common(1)[0][0],
                    "count": counts.most_common(1)[0][1],
                }
        return summary

    def _to_csv_bytes(self, rows: list[dict[str, Any]], dataset_id: str, schema_profile: dict[str, Any]) -> bytes:
        columns = [str(item["name"]) for item in schema_profile.get("columns", [])]
        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=[*columns, "dataset_id"], extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            normalized = {key: self._normalize_value(row.get(key, "")) for key in columns}
            normalized["dataset_id"] = dataset_id
            writer.writerow(normalized)
        return output.getvalue().encode("utf-8")

    def _ensure_glue_database(self) -> None:
        try:
            self.glue.get_database(Name=self.glue_database)
        except self.glue.exceptions.EntityNotFoundException:
            self.glue.create_database(DatabaseInput={"Name": self.glue_database})

    def _upsert_glue_table(self, dataset_id: str, schema_profile: dict[str, Any], data_key: str) -> None:
        table_name = self.get_table_name(dataset_id)
        columns = [
            {"Name": item["name"], "Type": self._athena_type(item["kind"])}
            for item in schema_profile.get("columns", [])
        ]
        columns.append({"Name": "dataset_id", "Type": "string"})
        table_input = {
            "Name": table_name,
            "TableType": "EXTERNAL_TABLE",
            "Parameters": {
                "classification": "csv",
                "EXTERNAL": "TRUE",
                "skip.header.line.count": "1",
            },
            "StorageDescriptor": {
                "Columns": columns,
                "Location": f"s3://{self.bucket_name}/{Path(data_key).parent.as_posix()}/",
                "InputFormat": "org.apache.hadoop.mapred.TextInputFormat",
                "OutputFormat": "org.apache.hadoop.hive.ql.io.HiveIgnoreKeyTextOutputFormat",
                "SerdeInfo": {
                    "SerializationLibrary": "org.apache.hadoop.hive.serde2.OpenCSVSerde",
                    "Parameters": {
                        "separatorChar": ",",
                        "quoteChar": "\"",
                        "escapeChar": "\\",
                    },
                },
            },
        }
        try:
            self.glue.get_table(DatabaseName=self.glue_database, Name=table_name)
            self.glue.update_table(DatabaseName=self.glue_database, TableInput=table_input)
        except self.glue.exceptions.EntityNotFoundException:
            self.glue.create_table(DatabaseName=self.glue_database, TableInput=table_input)

    def _write_dataset_manifest(self, dataset_id: str, source_name: str, schema_profile: dict[str, Any]) -> None:
        manifest = {
            "dataset_id": dataset_id,
            "source_name": source_name,
            "table_name": self.get_table_name(dataset_id),
            "updated_at": int(time.time()),
            "schema_columns": [item["name"] for item in schema_profile.get("columns", [])],
        }
        self.s3.put_object(
            Bucket=self.bucket_name,
            Key=f"datasets/{dataset_id}/manifest.json",
            Body=json.dumps(manifest, indent=2).encode("utf-8"),
            ContentType="application/json",
        )

    def _load_json(self, key: str) -> dict[str, Any] | None:
        try:
            response = self.s3.get_object(Bucket=self.bucket_name, Key=key)
        except ClientError:
            return None
        return json.loads(response["Body"].read().decode("utf-8"))

    def _list_common_prefixes(self, prefix: str) -> list[str]:
        paginator = self.s3.get_paginator("list_objects_v2")
        prefixes: list[str] = []
        for page in paginator.paginate(Bucket=self.bucket_name, Prefix=prefix, Delimiter="/"):
            prefixes.extend(item["Prefix"] for item in page.get("CommonPrefixes", []))
        return prefixes

    @staticmethod
    def _normalize_value(value: Any) -> Any:
        if isinstance(value, datetime):
            return value.isoformat()
        return value

    @staticmethod
    def _infer_kind(values: Iterable[Any], unique_values: list[str]) -> str:
        values = [value for value in values if value not in (None, "")]
        if not values:
            return "text"

        if all(AnalyticsStore._is_numeric(value) for value in values):
            return "numeric"
        if all(AnalyticsStore._is_datetime(value) for value in values):
            return "datetime"
        if len(unique_values) <= 50:
            return "categorical"
        return "text"

    @staticmethod
    def _is_numeric(value: Any) -> bool:
        if isinstance(value, (int, float)):
            return True
        try:
            float(str(value))
            return True
        except (TypeError, ValueError):
            return False

    @staticmethod
    def _is_datetime(value: Any) -> bool:
        if isinstance(value, datetime):
            return True
        try:
            datetime.fromisoformat(str(value).replace("Z", "+00:00"))
            return True
        except (TypeError, ValueError):
            return False

    @staticmethod
    def _athena_type(kind: str) -> str:
        return {
            "numeric": "double",
            "datetime": "string",
            "categorical": "string",
            "text": "string",
        }.get(kind, "string")
