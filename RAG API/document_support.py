from __future__ import annotations

import os
import csv
import json
from collections import Counter
from pathlib import Path
from typing import Iterable, List

from docx import Document as DocxDocument
from openpyxl import load_workbook
from PyPDF2 import PdfReader
from langchain_text_splitters import RecursiveCharacterTextSplitter


MAX_UPLOAD_BYTES = 5 * 1024 * 1024
EXCEPTION_UPLOAD_BYTES = 25 * 1024 * 1024
EXCEPTION_WORKSPACE = "test-big-001"
WORKSPACE_DOCUMENT_LIMIT = 10
PDF_PAGE_WARNING_THRESHOLD = 12
PDF_TEXT_ONLY_THRESHOLD = 20
PDF_PAGE_HARD_LIMIT = 35
ALLOWED_CONTENT_TYPES = {
    ".pdf": {"application/pdf"},
    ".txt": {"text/plain"},
    ".docx": {
        "application/vnd.openxmlformats-officedocument.wordprocessingml.document",
        "application/msword",
    },
    ".xlsx": {
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "application/vnd.ms-excel",
    },
}
TICKET_CONTENT_TYPES = {
    ".csv": {"text/csv", "application/csv", "application/vnd.ms-excel"},
    ".json": {"application/json", "text/json"},
}
GENERIC_AMBIGUOUS_TERMS = {
    "coverage",
    "claim",
    "claims",
    "policy",
    "benefit",
    "benefits",
    "premium",
    "resolution",
    "issue",
    "incident",
    "deductible",
}


def normalize_extension(filename: str) -> str:
    return Path(filename or "").suffix.lower()


def is_limit_exempt_workspace(index_name: str | None) -> bool:
    return (index_name or "").strip().lower() == EXCEPTION_WORKSPACE


def get_max_upload_bytes(index_name: str | None) -> int:
    return EXCEPTION_UPLOAD_BYTES if is_limit_exempt_workspace(index_name) else MAX_UPLOAD_BYTES


def get_upload_policy(index_name: str | None, existing_documents_count: int = 0) -> dict:
    exempt = is_limit_exempt_workspace(index_name)
    return {
        "workspace_id": (index_name or "").strip().lower(),
        "is_exception_workspace": exempt,
        "exception_workspace_id": EXCEPTION_WORKSPACE,
        "supported_types": ["pdf", "txt", "docx", "xlsx"],
        "max_upload_mb": round(get_max_upload_bytes(index_name) / (1024 * 1024)),
        "workspace_document_limit": None if exempt else WORKSPACE_DOCUMENT_LIMIT,
        "workspace_document_count": existing_documents_count,
        "pdf_page_warning_threshold": None if exempt else PDF_PAGE_WARNING_THRESHOLD,
        "pdf_text_only_threshold": None if exempt else PDF_TEXT_ONLY_THRESHOLD,
        "pdf_page_hard_limit": None if exempt else PDF_PAGE_HARD_LIMIT,
        "warnings": [
            "Only upload non-sensitive files. Do not upload critical, regulated, or PII data.",
            "Supported file types: PDF, TXT, DOCX, XLSX.",
            (
                f"Standard workspaces allow files up to {round(MAX_UPLOAD_BYTES / (1024 * 1024))} MB, "
                f"up to {WORKSPACE_DOCUMENT_LIMIT} indexed documents per workspace, and PDF guidance on page count."
                if not exempt
                else f"{EXCEPTION_WORKSPACE} is the test exception workspace. Higher upload limits and PDF limits are bypassed there."
            ),
            (
                f"PDFs over {PDF_TEXT_ONLY_THRESHOLD} pages switch to lower-cost text-only processing, and PDFs over "
                f"{PDF_PAGE_HARD_LIMIT} pages are blocked in standard workspaces."
                if not exempt
                else "Large PDFs in the exception workspace still upload at your own risk and may take longer to process."
            ),
        ],
    }


def get_pdf_page_count(file_path: str) -> int:
    reader = PdfReader(file_path)
    return len(reader.pages)


def validate_upload(
    filename: str,
    content_type: str,
    size_bytes: int,
    *,
    index_name: str | None = None,
    existing_documents_count: int | None = None,
    pdf_page_count: int | None = None,
) -> List[str]:
    extension = normalize_extension(filename)
    if extension not in ALLOWED_CONTENT_TYPES:
        raise ValueError(f"Unsupported file type '{extension or 'unknown'}'. Allowed: pdf, txt, docx, xlsx.")
    max_upload_bytes = get_max_upload_bytes(index_name)
    if size_bytes > max_upload_bytes:
        raise ValueError(f"File exceeds {round(max_upload_bytes / (1024 * 1024))} MB limit for this workspace.")
    allowed_content_types = ALLOWED_CONTENT_TYPES[extension]
    if content_type and content_type not in allowed_content_types:
        raise ValueError(f"Unexpected content type '{content_type}' for '{extension}'.")
    if is_limit_exempt_workspace(index_name):
        return []

    warnings: List[str] = []
    if existing_documents_count is not None and existing_documents_count >= WORKSPACE_DOCUMENT_LIMIT:
        raise ValueError(
            f"Workspace document limit reached. Standard workspaces allow up to {WORKSPACE_DOCUMENT_LIMIT} indexed documents."
        )
    if extension == ".pdf" and pdf_page_count is not None:
        if pdf_page_count > PDF_PAGE_HARD_LIMIT:
            raise ValueError(
                f"PDF exceeds the standard workspace page limit of {PDF_PAGE_HARD_LIMIT} pages. "
                f"Use '{EXCEPTION_WORKSPACE}' only if you intentionally need large-file testing."
            )
        if pdf_page_count > PDF_TEXT_ONLY_THRESHOLD:
            warnings.append(
                "Large PDF detected. To reduce demo cost, this upload will use text-only processing and skip image/table extraction."
            )
        elif pdf_page_count > PDF_PAGE_WARNING_THRESHOLD:
            warnings.append(
                "This PDF is larger than the recommended page count and may take longer to process."
            )
    return warnings


def infer_category(filename: str, sample_texts: Iterable[str]) -> str:
    combined = f"{filename}\n" + "\n".join(text for text in sample_texts if text)
    lowered = combined.lower()
    extension = normalize_extension(filename)
    scoring_rules = {
        "cloud_costing": {
            "keywords": (
                "aws",
                "amazon web services",
                "monthly cost",
                "monthly estimate",
                "cost estimate",
                "pricing",
                "budget",
                "forecast",
                "finops",
                "cloud cost",
                "service name",
                "estimated monthly",
                "ec2",
                "s3",
                "lambda",
                "cloudfront",
                "dynamodb",
                "fargate",
                "bedrock",
                "textract",
                "api gateway",
            ),
            "bonus": 0,
        },
        "support_tickets": {
            "keywords": (
                "ticket",
                "incident",
                "servicenow",
                "outage",
                "resolution",
                "short description",
                "assignment group",
                "opened by",
                "caller",
                "priority",
                "severity",
                "sla",
            ),
            "bonus": 0,
        },
        "medical_insurance": {
            "keywords": ("medical", "health", "hospital", "doctor", "patient", "clinic", "prescription"),
            "bonus": 0,
        },
        "auto_insurance": {
            "keywords": ("auto", "vehicle", "car", "driver", "collision", "accident", "garage"),
            "bonus": 0,
        },
        "travel_insurance": {
            "keywords": ("travel", "trip", "baggage", "flight", "journey", "passport", "hotel"),
            "bonus": 0,
        },
        "home_insurance": {
            "keywords": ("home", "property", "house", "dwelling", "contents", "fire", "theft"),
            "bonus": 0,
        },
        "insurance": {
            "keywords": ("insurance", "policy", "coverage", "claim", "premium", "insured", "benefit"),
            "bonus": 0,
        },
    }

    scores = {}
    for category, rule in scoring_rules.items():
        score = sum(1 for keyword in rule["keywords"] if keyword in lowered) + int(rule.get("bonus", 0))
        scores[category] = score

    insurance_categories = ("medical_insurance", "auto_insurance", "travel_insurance", "home_insurance", "insurance")
    insurance_signal = any(scores[category] > 0 for category in insurance_categories)

    if extension == ".pdf":
        scores["insurance"] += 2
    if extension == ".xlsx":
        scores["cloud_costing"] += 2
    if any(token in lowered for token in ("policy", "insurance", "coverage", "claim")):
        scores["insurance"] += 3
    if any(token in lowered for token in ("aws", "monthly cost", "pricing", "cost estimate", "service name", "estimated monthly")):
        scores["cloud_costing"] += 3
    if any(token in lowered for token in ("ticket", "servicenow", "assignment group", "short description")):
        scores["support_tickets"] += 3
    if insurance_signal and extension == ".pdf":
        scores["support_tickets"] = max(0, scores["support_tickets"] - 3)

    best_category = max(scores, key=scores.get)
    if scores[best_category] > 0:
        return best_category
    if extension == ".pdf":
        return "insurance"
    return "uncategorized"


def query_needs_clarification(query: str) -> bool:
    lowered = (query or "").lower()
    return any(term in lowered for term in GENERIC_AMBIGUOUS_TERMS)


def extract_text_chunks(file_path: str) -> List[str]:
    extension = normalize_extension(file_path)
    if extension == ".pdf":
        reader = PdfReader(file_path)
        text = "\n".join(page.extract_text() or "" for page in reader.pages)
        return _split_text(text)
    if extension == ".txt":
        text = Path(file_path).read_text(encoding="utf-8", errors="ignore")
        return _split_text(text)
    if extension == ".docx":
        document = DocxDocument(file_path)
        text = "\n".join(paragraph.text.strip() for paragraph in document.paragraphs if paragraph.text.strip())
        return _split_text(text)
    if extension == ".xlsx":
        workbook = load_workbook(file_path, read_only=True, data_only=True)
        chunks: List[str] = []
        for worksheet in workbook.worksheets:
            rows = []
            for row in worksheet.iter_rows(values_only=True):
                values = [str(cell).strip() for cell in row if cell not in (None, "")]
                if values:
                    rows.append(" | ".join(values))
            if rows:
                chunks.extend(_split_text(f"Sheet: {worksheet.title}\n" + "\n".join(rows)))
        return chunks
    raise ValueError(f"Text extraction not supported for '{extension}'.")


def build_text_result(chunks: List[str]) -> dict:
    return {
        str(index + 1): {
            "output": chunk,
            "page_numbers": ["1"],
            "url": [],
            "type": "TEXT",
        }
        for index, chunk in enumerate(chunks)
    }


def validate_ticket_upload(filename: str, content_type: str, size_bytes: int) -> None:
    extension = normalize_extension(filename)
    if extension not in TICKET_CONTENT_TYPES:
        raise ValueError("Unsupported ticket file type. Allowed: csv, json.")
    if size_bytes > MAX_UPLOAD_BYTES:
        raise ValueError("File exceeds 5 MB limit.")
    allowed_content_types = TICKET_CONTENT_TYPES[extension]
    if content_type and content_type not in allowed_content_types:
        raise ValueError(f"Unexpected content type '{content_type}' for '{extension}'.")


def extract_ticket_chunks(file_path: str) -> List[str]:
    extension = normalize_extension(file_path)
    if extension == ".csv":
        with open(file_path, "r", encoding="utf-8", errors="ignore", newline="") as handle:
            reader = csv.DictReader(handle)
            rows = [row for row in reader]
            return _build_ticket_summary_chunks(rows) + [
                "\n".join(
                    f"{key}: {value}"
                    for key, value in row.items()
                    if value not in (None, "")
                )
                for row in rows
            ]
    if extension == ".json":
        payload = json.loads(Path(file_path).read_text(encoding="utf-8", errors="ignore"))
        if isinstance(payload, dict):
            payload = payload.get("tickets", [payload])
        rows = [row for row in payload if isinstance(row, dict)]
        return _build_ticket_summary_chunks(rows) + [
            "\n".join(f"{key}: {value}" for key, value in row.items() if value not in (None, ""))
            for row in rows
        ]
    raise ValueError(f"Ticket extraction not supported for '{extension}'.")


def _split_text(text: str) -> List[str]:
    cleaned = (text or "").strip()
    if not cleaned:
        return []
    splitter = RecursiveCharacterTextSplitter(
        chunk_size=1200,
        chunk_overlap=150,
        separators=["\n\n", "\n", ". ", " ", ""],
    )
    return [chunk.strip() for chunk in splitter.split_text(cleaned) if chunk.strip()]


def _build_ticket_summary_chunks(rows: List[dict]) -> List[str]:
    if not rows:
        return []

    total_tickets = len(rows)
    category_counts = Counter((row.get("category") or "unknown").strip() for row in rows)
    priority_counts = Counter((row.get("priority") or "unknown").strip() for row in rows)
    status_counts = Counter((row.get("status") or "unknown").strip() for row in rows)
    source_counts = Counter((row.get("source") or "unknown").strip() for row in rows)
    assignment_counts = Counter((row.get("assignment_group") or "unknown").strip() for row in rows)
    summary_counts = Counter((row.get("summary") or "unknown").strip() for row in rows)

    chunks = [
        "\n".join(
            [
                f"ServiceNow ticket dataset overview",
                f"Total tickets: {total_tickets}",
                "Assignment groups by ticket volume:",
                _counter_to_markdown_table(assignment_counts, "assignment_group", "ticket_count"),
                "Categories by ticket volume:",
                _counter_to_markdown_table(category_counts, "category", "ticket_count"),
                "Priorities by ticket volume:",
                _counter_to_markdown_table(priority_counts, "priority", "ticket_count"),
            ]
        ),
        "\n".join(
            [
                "ServiceNow ticket operational summary",
                "Statuses by ticket volume:",
                _counter_to_markdown_table(status_counts, "status", "ticket_count"),
                "Sources by ticket volume:",
                _counter_to_markdown_table(source_counts, "source", "ticket_count"),
                "Most common ticket summaries:",
                _counter_to_markdown_table(summary_counts, "summary", "ticket_count"),
            ]
        ),
    ]
    return chunks


def _counter_to_markdown_table(counter: Counter, label: str, value_label: str, limit: int = 10) -> str:
    rows = [f"| {label} | {value_label} |", "| --- | ---: |"]
    for key, count in counter.most_common(limit):
        rows.append(f"| {key or 'unknown'} | {count} |")
    return "\n".join(rows)
